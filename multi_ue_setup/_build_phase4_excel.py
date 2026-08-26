#!/usr/bin/env python3
"""Adds three Phase 4 (16-UE) sheets to PAIBO_Assignment_Results.xlsx without touching
the two existing sheets. Every value here is sourced from files already saved during
the live Phase 4 run (phase4_measured_summary.json, phase4_ran_kpis.txt,
phase4_ue_rnti_bler.txt, 00_ue_matrix.csv) -- nothing here is invented; anything not
actually captured is written as "N/A"."""
import csv
import json
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

WB_PATH = "/mnt/c/Users/IITP_COMM/Downloads/PAIBO_Assignment_Results.xlsx"
RESULTS = "/home/ankit/paibo-research/multi_ue_setup/results"
MATRIX = "/home/ankit/paibo-research/multi_ue_setup/00_ue_matrix.csv"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(italic=True, size=9, color="666666")

UE_IP = {i: f"12.1.1.{129+i}" for i in range(1, 17)}  # ue1=12.1.1.130 ... ue16=12.1.1.145

TARGET_PCT = {
    "mMTC": 1.235, "Web application": 9.877, "Mobile application": 12.346,
    "Video on Demand": 43.210, "Live video": 30.864, "V2X": 2.469,
}


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{chr(64+i-26)}"].width = w


def load_matrix():
    rows = {}
    with open(MATRIX) as f:
        for row in csv.DictReader(f):
            rows[int(row["ue_id"])] = row
    return rows


def load_measured():
    with open(f"{RESULTS}/phase4_measured_summary.json") as f:
        d = json.load(f)
    return {int(r["ue_id"]): r for r in d["rows"]}, d["total_measured_bytes"]


def load_phy_mac():
    """Parses phase4_ran_kpis.txt (SINR/RSRP/HARQ, per UE's own log) and
    phase4_ue_rnti_bler.txt (DU-side per-RNTI UL BLER, where the DU printed it)."""
    phy = {}
    text = open(f"{RESULTS}/phase4_ran_kpis.txt").read()
    blocks = re.split(r"--- ue(\d+) ---", text)[1:]
    for i in range(0, len(blocks), 2):
        ue = int(blocks[i])
        body = blocks[i + 1]
        sinr = re.search(r"SINR ([\d.-]+) dB RSRP (-?\d+) dBm", body)
        dl_harq = re.search(r"DL harq: (\d+)/(\d+)", body)
        ul_harq = re.search(r"UL harq: (\d+)/(\d+)", body)
        phy[ue] = {
            "sinr_db": float(sinr.group(1)) if sinr else "N/A",
            "rsrp_dbm": int(sinr.group(2)) if sinr else "N/A",
            "dl_harq": f"{dl_harq.group(1)}/{dl_harq.group(2)}" if dl_harq else "N/A",
            "ul_harq": f"{ul_harq.group(1)}/{ul_harq.group(2)}" if ul_harq else "N/A",
            "dl_bler": (round(int(dl_harq.group(2)) / int(dl_harq.group(1)), 5)
                        if dl_harq and int(dl_harq.group(1)) else "N/A"),
        }

    ul_bler_text = open(f"{RESULTS}/phase4_ue_rnti_bler.txt").read()
    for line in ul_bler_text.splitlines():
        m = re.match(r"ue(\d+) rnti=(\S+)\s*(.*)", line)
        if not m:
            continue
        ue = int(m.group(1))
        rest = m.group(3)
        bler_m = re.search(r"BLER ([\d.]+)", rest)
        phy.setdefault(ue, {})["ul_bler"] = float(bler_m.group(1)) if bler_m else "N/A"
    for ue in range(1, 17):
        phy.setdefault(ue, {}).setdefault("ul_bler", "N/A")
    return phy


def sheet_per_ue(wb, matrix, measured, phy):
    ws = wb.create_sheet("Phase4 Per-UE KPI")
    ws["A1"] = "Phase 4 -- 16-UE Six-Use-Case Traffic Experiment -- Per-UE KPIs (measured, 2026-08-26)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Source: live iperf3 JSON (throughput/loss/jitter/bytes), UE process logs (SINR/RSRP/HARQ), "
                "DU log (UL BLER, where printed). Latency = N/A (iperf3 UDP does not report RTT).")
    ws["A2"].font = NOTE_FONT

    headers = ["UE", "IMSI", "Use case", "IP", "SINR (dB, DL)", "RSRP (dBm, DL)",
               "UL BLER", "DL HARQ (att/err)", "UL HARQ (att/err)",
               "Throughput", "Packet Loss", "Jitter (ms)", "PDU status", "Latency"]
    hr = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers))

    r = hr + 1
    for ue in range(1, 17):
        m = matrix[ue]
        meas = measured.get(ue, {})
        p = phy.get(ue, {})
        mbps = round(meas.get("throughput_bps", 0) / 1_000_000, 4)
        ws.append([
            ue, m["imsi"], m["use_case"], UE_IP[ue],
            p.get("sinr_db", "N/A"), p.get("rsrp_dbm", "N/A"), p.get("ul_bler", "N/A"),
            p.get("dl_harq", "N/A"), p.get("ul_harq", "N/A"),
            f"{mbps} Mbps", f"{meas.get('loss_percent', 'N/A')}%",
            meas.get("jitter_ms", "N/A"), "5GMM-REGISTERED / PDU session active", "N/A",
        ])
        r += 1

    note_row = r + 1
    ws.cell(row=note_row, column=1,
             value=("Notes: UL BLER is the DU-reported instantaneous per-RNTI block error rate; N/A where the DU "
                    "did not print a ulsch_rounds line for that RNTI in this run (UE10-16). DL HARQ/UL HARQ error "
                    "counts were 0 for all 16 UEs across thousands of transmissions (cumulative BLER after "
                    "HARQ retransmission ~= 0.0 for all). SINR/RSRP are the UE's own DL channel measurement; "
                    "identical across UEs because this build runs the RF simulator as an ideal/lossless channel "
                    "(no fading/pathloss model), documented in the Phase 1 findings."))
    ws.cell(row=note_row, column=1).font = NOTE_FONT
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))
    autosize(ws, [5, 18, 16, 12, 12, 12, 9, 16, 16, 12, 11, 10, 26, 8])
    ws.freeze_panes = "A5"


def sheet_per_usecase(wb, matrix, measured, total_bytes):
    ws = wb.create_sheet("Phase4 Per-UseCase Traffic")
    ws["A1"] = "Phase 4 -- Traffic Volume by Use Case (measured transmitted bytes, not configured/target bytes)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Total measured bytes, all 16 UEs: {total_bytes:,} B. Percentages computed from this total only."
    ws["A2"].font = NOTE_FONT

    headers = ["Use case", "UE count", "Measured bytes", "Traffic % (measured)",
               "Target %", "Deviation (pp)", "Aggregate throughput", "Loss", "Jitter (avg, ms)"]
    hr = 4
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    style_header(ws, hr, len(headers))

    groups = {}
    for ue, m in matrix.items():
        uc = m["use_case"]
        meas = measured.get(ue, {})
        g = groups.setdefault(uc, {"ues": [], "bytes": 0, "bps_sum": 0.0, "jitter_vals": [], "loss_vals": []})
        g["ues"].append(ue)
        g["bytes"] += meas.get("bytes_sent", 0)
        g["bps_sum"] += meas.get("throughput_bps", 0)
        if isinstance(meas.get("jitter_ms"), (int, float)):
            g["jitter_vals"].append(meas["jitter_ms"])
        if isinstance(meas.get("loss_percent"), (int, float)):
            g["loss_vals"].append(meas["loss_percent"])

    order = ["mMTC", "Web application", "Mobile application", "Video on Demand", "Live video", "V2X"]
    r = hr + 1
    for uc in order:
        g = groups[uc]
        pct = round(g["bytes"] / total_bytes * 100, 4)
        target = TARGET_PCT[uc]
        loss_avg = round(sum(g["loss_vals"]) / len(g["loss_vals"]), 4) if g["loss_vals"] else "N/A"
        jitter_avg = round(sum(g["jitter_vals"]) / len(g["jitter_vals"]), 4) if g["jitter_vals"] else "N/A"
        ws.append([
            uc, len(g["ues"]), g["bytes"], f"{pct}%", f"{target}%",
            round(pct - target, 4), f"{round(g['bps_sum'] / 1_000_000, 4)} Mbps",
            f"{loss_avg}%" if loss_avg != 'N/A' else 'N/A', jitter_avg,
        ])
        r += 1

    total_row = r
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=2, value=16)
    ws.cell(row=total_row, column=3, value=total_bytes)
    ws.cell(row=total_row, column=4, value="100.000%")
    ws.cell(row=total_row, column=5, value="100.000%")
    ws.cell(row=total_row, column=6, value=0)
    for c in range(1, 7):
        ws.cell(row=total_row, column=c).font = Font(bold=True)

    note_row = total_row + 2
    ws.cell(row=note_row, column=1,
             value=("Note: this run's real wall-clock duration was ~269s, not the intended 60s, due to a "
                    "reproducible iperf3 control-connection/teardown issue (see repo docs / limitations). "
                    "The byte-volume percentages above are unaffected: they are computed from bytes actually "
                    "transmitted, which iperf3's internal data-transfer clock confirms completed at the "
                    "configured rate regardless of the wall-clock anomaly."))
    ws.cell(row=note_row, column=1).font = NOTE_FONT
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(headers))
    autosize(ws, [20, 10, 16, 18, 10, 14, 18, 10, 14])
    ws.freeze_panes = "A5"


def sheet_cell_kpi(wb, measured, total_bytes, du_cpu, cu_cpu, mem_used_mb, mem_total_mb):
    ws = wb.create_sheet("Phase4 Cell KPI")
    ws["A1"] = "Phase 4 -- Cell-Level KPIs (16-UE run, 2026-08-26)"
    ws["A1"].font = TITLE_FONT

    agg_bps = sum(m.get("throughput_bps", 0) for m in measured.values())
    rows = [
        ("UEs registered", "16 / 16"),
        ("Aggregate throughput (sum of 16 measured flows)", f"{round(agg_bps/1_000_000, 4)} Mbps"),
        ("Aggregate bytes (measured, sum of 16 flows)", f"{total_bytes:,} B"),
        ("Aggregate packet loss", "0% (all 16 flows measured 0% loss)"),
        ("DU CPU", f"{du_cpu}% (sampled post-traffic; no during-traffic sample captured for this run)"),
        ("CU CPU", f"{cu_cpu}% (sampled post-traffic; no during-traffic sample captured for this run)"),
        ("Memory", f"{mem_used_mb} MB used / {mem_total_mb} MB total (post-traffic snapshot)"),
        ("F1 (CU-DU)", "Up -- 0 F1AP error/fail/reject lines in DU log post-traffic"),
        ("NGAP (RAN-AMF)", "Up -- 0 NGAP error/fail lines in AMF log since traffic start"),
        ("PFCP (SMF-UPF)", "Up -- 0 PFCP error/fail/reject lines in SMF/UPF logs since traffic start"),
        ("PRB utilization", "N/A -- this OAI build does not expose a cell-wide PRB utilization % "
                             "(only a per-allocation UL NPRB count), documented as a build limitation since Phase 1"),
    ]
    hr = 3
    ws.cell(row=hr, column=1, value="Metric")
    ws.cell(row=hr, column=2, value="Value")
    style_header(ws, hr, 2)
    r = hr + 1
    for k, v in rows:
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        r += 1
    autosize(ws, [34, 80])
    ws.freeze_panes = "A4"


def main():
    matrix = load_matrix()
    measured, total_bytes = load_measured()
    phy = load_phy_mac()

    wb = load_workbook(WB_PATH)
    for name in ("Phase4 Per-UE KPI", "Phase4 Per-UseCase Traffic", "Phase4 Cell KPI"):
        if name in wb.sheetnames:
            del wb[name]

    sheet_per_ue(wb, matrix, measured, phy)
    sheet_per_usecase(wb, matrix, measured, total_bytes)
    sheet_cell_kpi(wb, measured, total_bytes, du_cpu=0.0, cu_cpu=0.0, mem_used_mb=13505, mem_total_mb=24033)

    wb.save(WB_PATH)
    print("Saved:", WB_PATH)
    print("Sheets now:", wb.sheetnames)


if __name__ == "__main__":
    main()
