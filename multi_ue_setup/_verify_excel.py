from openpyxl import load_workbook

wb = load_workbook("/mnt/c/Users/IITP_COMM/Downloads/PAIBO_Assignment_Results.xlsx")
print("Sheets:", wb.sheetnames)
for name in ["Phase4 Per-UE KPI", "Phase4 Per-UseCase Traffic", "Phase4 Cell KPI"]:
    ws = wb[name]
    print(f"\n===== {name} (dims={ws.dimensions}) =====")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True):
        print(row)
